#encoding=utf-8
from openpyxl import load_workbook
from openpyxl.styles import Font,colors
import locale,time

class ParseExcel(object):
    def __init__(self,excel_file_path):
        self.excel_file_path = excel_file_path
        self.wb = load_workbook(excel_file_path)
        #获取当前的第一个表格
        self.sheet = self.wb[self.wb.sheetnames[0]]
        self.style_dict = {"red" : colors.RED,"green" : colors.GREEN}

    #获取所有的行
    def get_all_rows(self,sheet_name = None):
        if sheet_name in self.wb.sheetnames:
            result = self.set_sheet_by_name(sheet_name)
            if result == False:
                return None
        return list(self.sheet.rows)#返回列表

    #获取所有的sheet名字
    def get_all_sheet_names(self):
        return  self.wb.sheetnames

    #通过索引获取当前sheet名字
    def get_sheet_name_by_index(self,index):
        return self.wb.sheetnames[index - 1]

    #获取测试文件路径
    def get_excel_file_path(self):
        return self.excel_file_path

    #创建sheet表
    def create_sheet(self,sheet_name,position=None):
        try:
            if position:
                self.wb.create_sheet(sheet_name,position)
            else:
                self.wb.create_sheet(sheet_name)
            self.wb.save(self.excel_file_path)
            return True
        except Exception as e:
            print(e)
            return False

    #根据sheet名设置当前sheet
    def set_sheet_by_name(self,sheet_name):
        if sheet_name in self.wb.sheetnames:
            self.sheet = self.wb[sheet_name]
            return True
        else:
            print("sheet [%s] 不存在" %sheet_name)
            return False

    #根据index设置当前sheet
    def set_sheet_by_index(self,index):
        try:
            self.sheet = self.wb[self.wb.sheetnames[index]]
        except Exception as e:
            print(e)

    #读取某个sheet某个单元格的值
    def get_cell_value(self,row_no,col_no,sheet_name=None):
        if sheet_name in self.wb.sheetnames:
            result = self.set_sheet_by_name(sheet_name)
            if result == False:
                return None
        return self.sheet.cell(row_no,col_no).value

    #读取一行的值
    def get_row_vaules(self,row_no,sheet_name=None):
        cell_values = []
        if sheet_name in self.wb.sheetnames:
            result = self.set_sheet_by_name(sheet_name)
            if not result:
                return None
        row = list(self.sheet.rows)[row_no - 1]
        for cell  in row:
            cell_values.append(cell.value)
        return cell_values

    #读取某个sheet中所有行的单元格内容，存入二维列表
    def get_all_rows_values(self,sheet_name = None):
        all_cell_values = []
        #如果指定了sheet名，先设置当前sheet
        if sheet_name in self.wb.sheetnames:
            result = self.set_sheet_by_name(sheet_name)
            if not result:
                return None
        for row in list(self.sheet.rows):
            row_values = []
            for cell in row:
                row_values.append(cell.value)
            all_cell_values.append(row_values)
        return all_cell_values

    #读取某一列的值
    def get_col_values(self,col_no,sheet_name=None):
        col_values = []
        if sheet_name in self.wb.sheetnames:
            result = self.set_sheet_by_name(sheet_name)
            if not result:
                return None

        col = list(self.sheet.columns)[col_no - 1]
        for cell in col:
            col_values.append(cell.value)
        return col_values

    #读取某个范围单元格的值
    def get_some_values(self,min_row_no,min_col_no,max_row_no,max_col_no,sheet_name = None):
        if sheet_name in self.wb.sheetnames:
            result = self.set_sheet_by_name(sheet_name)
            if not result:
                return None
        values = []#二维列表，存储所有单元格的数据
        for r in range(min_row_no,max_row_no + 1):
            row_values = []
            for c in range(min_col_no,max_col_no + 1):
                row_values.append(self.sheet.cell(row=r,column=c).value)
            values.append(row_values)
        return values

    #写单元格内容
    def write_cell_value(self,row_no,col_no,content,style=None,sheet_name=None):
        if sheet_name in self.wb.sheetnames:
            result = self.set_sheet_by_name(sheet_name)
            if not result:
                return None
        self.sheet.cell(row_no,col_no).value = content
        if style:
            self.sheet.cell(row_no,col_no).font = Font(color=self.style_dict[style])
        self.wb.save(self.excel_file_path)
        return True

    #写单元格时间
    def write_current_time(self,row_no,col_no,style=None,sheet_name = None):
        if sheet_name in self.wb.sheetnames:
            result = self.set_sheet_by_name(sheet_name)
            if not result:
                return None
        self.sheet.cell(row_no,col_no).value = time.strftime("%Y-%m-%d %H:%M:%S")
        if style:
            self.sheet.cell(row_no,col_no).font = Font(color=self.style_dict[style])
        self.wb.save(self.excel_file_path)

if __name__ == "__main__":
    pe = ParseExcel("d:\\接口测试数据.xlsx")
    # print(pe.get_excel_file_path())
    # print(pe.get_all_rows())
    print(pe.get_all_rows_values())
    print(pe.get_all_sheet_names())
    print(pe.get_sheet_name_by_index(1))
    print(pe.get_some_values(1,1,3,3))
    print(pe.get_row_vaules(1))
    print(pe.get_col_values(1))
    print(pe.get_cell_value(1,2))
    pe.set_sheet_by_name("注册登录")
    print(pe.get_all_rows())
    print(pe.get_all_rows_values())
    pe.create_sheet("tttt")
    pe.write_cell_value(5,5,100,style="red")
    pe.write_current_time(5,6,style="green")